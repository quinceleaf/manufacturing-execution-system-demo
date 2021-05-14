# TODO: material_calculate_cost_cumulative()
# TODO: material_calculate_cost_direct()
# TODO: material_calculate_cost_exp()
# TODO: material_calculate_cost_ma3()
# TODO: material_calculate_cost_ma5()
# TODO: material_calculate_cost_ma7()
# TODO: material_calculate_cost_naive()


# TODO: product_calculate_cost_cumulative()
# TODO: product_calculate_cost_direct()
# TODO: product_calculate_cost_exp()
# TODO: product_calculate_cost_ma3()
# TODO: product_calculate_cost_ma5()
# TODO: product_calculate_cost_ma7()
# TODO: product_calculate_cost_naive()

# --- DJANGO IMPORTS
from django.db.models import Model, QuerySet
from django.http import HttpResponse
from django.template.loader import get_template, render_to_string
from django.utils import timezone


# --- PYTHON UTILITY IMPORTS
from collections import deque
import csv
import datetime
import datetime as dt
import decimal
from decimal import Decimal as D
import json
import math
import re
from tempfile import NamedTemporaryFile
from typing import Tuple


# --- THIRD-PARTY IMPORTS
from openpyxl import load_workbook, Workbook
from openpyxl.utils import absolute_coordinate, quote_sheetname
from openpyxl.worksheet.cell_range import CellRange
from openpyxl.worksheet.datavalidation import DataValidation
import pytz
from weasyprint import HTML, CSS
from weasyprint.fonts import FontConfiguration


# --- APPLICATION IMPORTS
from apps.masterdata import models, selectors
from apps.common.converters import DecimalEncoder
from apps.common import services as common_services


# --- PARAMETERS
EASTERN_TZ = pytz.timezone("US/Eastern")


# –––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––
# UNITS OF MEASUREMENT
# –––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––


STANDARD_UNIT_TYPES = ["WEIGHT", "VOLUME", "EACH"]

STANDARD_UNITS = {"WEIGHT": "lb", "VOLUME": "fl oz", "EACH": "each"}

CONVERSION_DICT = {
    # VOLUMES
    "c,c": "1.00000000",
    "c,fl oz": "8.00000000",
    "c,gal": "0.03125000",
    "c,L": "0.24",
    "c,ml": "239.999808000154",
    "c,pt": "0.50000000",
    "c,qt": "0.12500000",
    "c,Tbsp": "16.00000000",
    "c,tsp": "48.00000000",
    "fl oz,c": "0.12500000",
    "fl oz,fl oz": "1.00000000",
    "fl oz,gal": "0.00781250",
    "fl oz,L": "0.0295735",
    "fl oz,ml": "29.5735494174011",
    "fl oz,pt": "0.06250000",
    "fl oz,qt": "0.03125000",
    "fl oz,Tbsp": "2.00000000",
    "fl oz,tsp": "6.00000000",
    "gal,c": "32.00000000",
    "gal,fl oz": "128.00000000",
    "gal,gal": "1.00000000",
    "gal,L": "3.78541",
    "gal,ml": "3785.41253425798",
    "gal,pt": "8.00000000",
    "gal,qt": "4.00000000",
    "gal,Tbsp": "256.00000000",
    "gal,tsp": "768.00000000",
    "L,c": "4.16666666666667",
    "L,fl oz": "33.8140565032884",
    "L,gal": "0.264172176857989",
    "L,L": "1",
    "L,L": "1",
    "L,ml": "0.001",
    "L,ml": "1000",
    "L,pt": "2.11337853145553",
    "L,qt": "1.05668814913674",
    "L,Tbsp": "67.6278843292666",
    "L,tsp": "202.884201812973",
    "ml,c": "0.00416667",
    "ml,fl oz": "0.033814",
    "ml,gal": "0.000264172",
    "ml,L": "0.001",
    "ml,L": "1000",
    "ml,ml": "1",
    "ml,ml": "1",
    "ml,pt": "0.00211338",
    "ml,qt": "0.00105669",
    "ml,Tbsp": "0.067628",
    "ml,tsp": "0.202884",
    "pt,c": "2.00000000",
    "pt,fl oz": "16.00000000",
    "pt,gal": "0.12500000",
    "pt,L": "0.473176",
    "pt,ml": "473.17567119969",
    "pt,pt": "1.00000000",
    "pt,qt": "0.50000000",
    "pt,Tbsp": "32.00000000",
    "pt,tsp": "96.00000000",
    "qt,c": "8.00000000",
    "qt,fl oz": "32.00000000",
    "qt,gal": "0.25000000",
    "qt,L": "0.946353",
    "qt,ml": "946.351342399379",
    "qt,pt": "2.00000000",
    "qt,qt": "1.00000000",
    "qt,Tbsp": "64.00000000",
    "qt,tsp": "192.00000000",
    "Tbsp,c": "0.0625",
    "Tbsp,fl oz": "0.5",
    "Tbsp,gal": "0.00390625",
    "Tbsp,L": "0.0147868",
    "Tbsp,ml": "14.7867747087005",
    "Tbsp,pt": "0.03125",
    "Tbsp,qt": "0.015625",
    "Tbsp,Tbsp": "1",
    "Tbsp,tsp": "3",
    "tsp,c": "0.020833333",
    "tsp,fl oz": "0.166666667",
    "tsp,gal": "0.001302083",
    "tsp,L": "0.00492892",
    "tsp,ml": "4.92892490290018",
    "tsp,pt": "0.010416667",
    "tsp,qt": "0.005208333",
    "tsp,Tbsp": "0.333333333",
    "tsp,tsp": "1",
    # WEIGHTS
    "kg,kg": "1",
    "kg,g": "1000",
    "kg,lb": "2.202643172",
    "kg,oz": "35.24229075",
    "g,kg": "0.001",
    "g,g": "1",
    "g,lb": "0.002202643",
    "g,oz": "0.035242291",
    "lb,kg": "0.454",
    "lb,g": "454",
    "lb,lb": "1",
    "lb,oz": "16",
    "oz,kg": "0.028375",
    "oz,g": "28.375",
    "oz,lb": "0.0625",
    "oz,oz": "1",
}

""" Units are grouped by their system and type, then ordered from smallest to largest """
""" Ladder is referenced to determine the appropriate unit to attempt to convert to for display """
UNIT_LADDER = {
    "METRIC_WEIGHT": ["g", "kg"],
    "METRIC_VOLUME": ["ml", "L"],
    "US_WEIGHT": ["oz", "lb"],
    "US_VOLUME": ["tsp", "Tbsp", "fl oz", "c", "pt", "qt", "gal"],
}


def convert_units(quantity, input_unit_symbol, output_unit_symbol):
    evaluate = f"{input_unit_symbol},{output_unit_symbol}"
    ratio = CONVERSION_DICT[evaluate]
    output_quantity = k = D(ratio) * D(quantity)
    return D(output_quantity).quantize(D("0.001"))


def standardize_units(
    *, quantity: decimal.Decimal, unit: models.UnitMeasurement
) -> Tuple[decimal.Decimal, models.UnitMeasurement]:
    """ given a quantity & unit for a measurement, convert to the unit specified as standard by user for that unit_type """
    if unit.unit_type not in STANDARD_UNIT_TYPES:
        return (quantity, unit)

    if unit.unit_type == "EACH":
        standardized_quantity = quantity
        standardized_unit = models.UnitMeasurement.objects.get(symbol="ea")
    else:
        if unit.unit_type == "WEIGHT":
            # check for default set by tenant
            if models.Settings.default_unit_weight:
                standardized_unit = models.Settings.default_unit_weight
                # else default to lbs
                standardized_unit = models.UnitMeasurement.objects.get(symbol="lb")
        else:
            # check for default set by tenant
            if models.Settings.default_unit_volume:
                standardized_unit = models.Settings.default_unit_volume
                # else default to fl oz
                standardized_unit = models.UnitMeasurement.objects.get(symbol="fl oz")
        standardized_quantity = convert_units(
            D(quantity),
            unit.symbol,
            standardized_unit.symbol,
        )

    return (standardized_quantity, standardized_unit)


def adjust_units_for_display(quantity, unit):
    """
    Scales quantities/units as appropriate for more convenient measurement
    so that a scaled BOM doesn't present a line specifying 160 tsp of item, for example

    if unit_type == EACHES: continue, non-applicable

    for WEIGHT and VOLUME, have metric and US customary ladders
    lookup which a unit belongs to, then scale it up/down ladder
    """
    from apps.masterdata.models import UnitMeasurement

    """ return immediately if not a scalable unit type """
    SCALABLE_UNITS = ["WEIGHT", "VOLUME"]
    if unit.unit_type not in SCALABLE_UNITS:
        return quantity, unit

    """ return immediately if not unit included in the ladders """
    if unit.symbol not in UNIT_LADDER.get(f"{unit.unit_system}_{unit.unit_type}", None):
        return quantity, unit

    ladder_branch = UNIT_LADDER[f"{unit.unit_system}_{unit.unit_type}"]
    current_quantity = D(quantity)
    current_unit = unit
    seeking = True

    """ success condition:
    if current_unit.display_quantity_smallest <=current_quantity <= current_unit.display_quantity_largest
    """

    while seeking:

        if (
            current_unit.display_quantity_smallest
            <= current_quantity
            <= current_unit.display_quantity_largest
        ):
            """ success condition met """
            break

        current_rank = ladder_branch.index(current_unit.symbol)

        if (
            current_quantity > current_unit.display_quantity_largest
            and current_rank != len(ladder_branch) - 1
        ):
            output_unit_symbol = ladder_branch[current_rank + 1]
        elif (
            current_quantity < current_unit.display_quantity_smallest
            and current_rank != 0
        ):
            output_unit_symbol = ladder_branch[current_rank - 1]
        else:
            return round(current_quantity, 3), current_unit

        current_quantity = convert_units(
            current_quantity, current_unit.symbol, output_unit_symbol
        )
        current_unit = UnitMeasurement.objects.get(symbol=output_unit_symbol)

    return round(current_quantity, 3), current_unit


# –––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––
# BILL OF MATERIALS
# –––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––


def bill_of_materials_create(
    *, product: models.Product, team: models.Team
) -> models.BillOfMaterials:
    bill_of_materials = models.BillOfMaterials(product=product, team=team)
    bill_of_materials.full_clean()
    bill_of_materials.save()

    return bill_of_materials


def bill_of_materials_characteristics_create(
    *,
    leadtime: int = None,
    temperature_preparation: str = None,
    temperature_storage: str = None,
    temperature_service: str = None,
    note_production: str = None,
    total_active_time: decimal.Decimal = D(0),
    total_inactive_time: decimal.Decimal = D(0),
    staff_count: int = 0,
    note_labor: str = None,
    bill_of_materials: models.BillOfMaterials,
) -> models.BillOfMaterialsCharacteristics:
    characteristics = models.BillOfMaterialsCharacteristics(
        leadtime=leadtime,
        temperature_preparation=temperature_preparation,
        temperature_storage=temperature_storage,
        temperature_service=temperature_service,
        note_production=note_production,
        total_active_time=total_active_time,
        total_inactive_time=total_inactive_time,
        staff_count=staff_count,
        note_labor=note_labor,
        bill_of_materials=bill_of_materials,
    )
    characteristics.full_clean()
    characteristics.save()

    return characteristics


def bill_of_materials_line_create(
    *,
    sequence: int,
    quantity: decimal.Decimal,
    quantity_standard: decimal.Decimal = D(0),
    note: str = None,
    bill_of_materials: models.BillOfMaterials,
    item: models.Item,
    unit: models.UnitMeasurement,
    unit_standard: models.UnitMeasurement = None,
) -> models.BillOfMaterialsLine:

    quantity_standard, unit_standard = standardize_units(quantity=quantity, unit=unit)

    line = models.BillOfMaterialsLine(
        sequence=sequence,
        quantity=quantity,
        quantity_standard=quantity_standard,
        unit=unit,
        unit_standard=unit_standard,
        item=item,
        note=note,
        bill_of_materials=bill_of_materials,
    )
    line.full_clean()
    line.save()

    return line


def bill_of_materials_line_update(
    *, bill_of_materials_line_id: str, data
) -> models.BillOfMaterialsLine:

    # TODO: Check permissions

    line = selectors.bill_of_materials_line_detail(
        bill_of_materials_line_id=bill_of_materials_line_id
    )

    valid_fields = ["sequence", "quantity", "note", "unit", "item"]

    for field in valid_fields:
        if field in data:
            setattr(some_object, field, data[field])

    if data["quantity"] and data["unit"]:
        data["quantity_standard"], data["unit_standard"] = standardize_units(
            quantity=data["quantity"], unit=data["unit"]
        )

    line.full_clean()
    line.save()

    return line


def bill_of_materials_line_update_standardize(
    *,
    bill_of_materials_line: models.BillOfMaterialsLine,
    quantity: decimal.Decimal,
    unit: models.UnitMeasurement,
) -> models.BillOfMaterialsLine:

    # TODO: Check permissions

    # line = selectors.bill_of_materials_line_detail(
    #     bill_of_materials_line_id=bill_of_materials_line_id
    # )

    if quantity and unit:
        quantity_standard, unit_standard = standardize_units(
            quantity=quantity, unit=unit
        )

    bill_of_materials_line.quantity_standard = quantity_standard
    bill_of_materials_line.unit_standard = unit_standard

    bill_of_materials_line.full_clean()
    bill_of_materials_line.save()

    return bill_of_materials_line


def bill_of_materials_note_create(
    *, note: str, bill_of_materials: models.BillOfMaterials
) -> models.BillOfMaterialsNote:
    bill_of_materials_note = models.BillOfMaterialsNote(
        note=note, bill_of_materials=bill_of_materials
    )
    bill_of_materials_note.full_clean()
    bill_of_materials_note.save()

    return bill_of_materials_note


def bill_of_materials_note_update(
    *, bill_of_materials_note_id: str, data
) -> models.BillOfMaterialsNote:

    # TODO: Check permissions

    bill_of_materials_note = selectors.bill_of_materials_note_detail(
        bill_of_materials_note_id=bill_of_materials_note_id
    )

    valid_fields = [
        "note",
    ]

    for field in valid_fields:
        if field in data:
            setattr(some_object, field, data[field])

    bill_of_materials_note.full_clean()
    bill_of_materials_note.save()

    return line


def bill_of_materials_procedure_create(
    *, language: str, procedure: str, bill_of_materials: models.BillOfMaterials
) -> models.BillOfMaterialsProcedure:
    bill_of_materials_procedure = models.BillOfMaterialsProcedure(
        language=language,
        procedure=procedure,
        bill_of_materials=bill_of_materials,
    )
    bill_of_materials_procedure.full_clean()
    bill_of_materials_procedure.save()

    return bill_of_materials_procedure


def bill_of_materials_procedure_update(
    *, bill_of_materials_procedure_id: str, data
) -> models.BillOfMaterialsProcedure:

    # TODO: Check permissions

    bill_of_materials_procedure = selectors.bill_of_materials_procedure_detail(
        bill_of_materials_procedure_id=bill_of_materials_procedure_id
    )

    valid_fields = ["language", "procedure"]

    for field in valid_fields:
        if field in data:
            setattr(some_object, field, data[field])

    bill_of_materials_procedure.full_clean()
    bill_of_materials_procedure.save()

    return bill_of_materials_procedure


def bill_of_materials_resource_create(
    *,
    sequence: int,
    capacity_required: decimal.Decimal,
    changeover_required: decimal.Decimal = D(0),
    resource: models.Resource,
    note: str,
    bill_of_materials: models.BillOfMaterials,
) -> models.BillOfMaterialsResource:
    bill_of_materials_resource = models.BillOfMaterialsResource(
        sequence=sequence,
        capacity_required=capacity_required,
        changeover_required=changeover_required,
        resource=resource,
        note=note,
        bill_of_materials=bill_of_materials,
    )
    bill_of_materials_resource.full_clean()
    bill_of_materials_resource.save()

    return bill_of_materials_resource


def bill_of_materials_resource_update(
    *, bill_of_materials_resource_id: str, data
) -> models.BillOfMaterialsResource:

    # TODO: Check permissions

    bill_of_materials_resource = selectors.bill_of_materials_resource_detail(
        bill_of_materials_resource_id=bill_of_materials_resource_id
    )

    valid_fields = [
        "sequence",
        "capacity_required",
        "changeover_required",
        "note",
        "resource",
    ]

    for field in valid_fields:
        if field in data:
            setattr(some_object, field, data[field])

    bill_of_materials_resource.full_clean()
    bill_of_materials_resource.save()

    return bill_of_materials_resource


def bill_of_materials_yield_create(
    *,
    quantity_weight: decimal.Decimal = None,
    unit_weight: models.UnitMeasurement = None,
    quantity_weight_standard: decimal.Decimal = None,
    unit_weight_standard: models.UnitMeasurement = None,
    quantity_volume: decimal.Decimal = None,
    unit_volume: models.UnitMeasurement = None,
    quantity_volume_standard: decimal.Decimal = None,
    unit_volume_standard: models.UnitMeasurement = None,
    quantity_each: decimal.Decimal = None,
    note_each: str = None,
    unit_each: models.UnitMeasurement = None,
    scale_multiple_smallest: decimal.Decimal = None,
    scale_multiple_largest: decimal.Decimal = None,
    bill_of_materials: models.BillOfMaterials,
) -> models.BillOfMaterialsYield:

    if quantity_weight and unit_weight:
        quantity_weight_standard, unit_weight_standard = standardize_units(
            quantity=quantity_weight, unit=unit_weight
        )

    if quantity_volume and unit_volume:
        quantity_volume_standard, unit_volume_standard = standardize_units(
            quantity=quantity_volume, unit=unit_volume
        )

    yields = models.BillOfMaterialsYield(
        quantity_weight=quantity_weight,
        unit_weight=unit_weight,
        quantity_volume=quantity_volume,
        unit_volume=unit_volume,
        quantity_each=quantity_each,
        unit_each=unit_each,
        note_each=note_each,
        scale_multiple_smallest=scale_multiple_smallest,
        scale_multiple_largest=scale_multiple_largest,
        bill_of_materials=bill_of_materials,
    )
    yields.full_clean()
    yields.save()

    return yields


def bill_of_materials_yield_update(
    *, bill_of_materials_yield_id: str, data
) -> models.BillOfMaterialsYield:

    # TODO: Check permissions

    yields = selectors.bill_of_materials_yield_detail(
        bill_of_materials_yield_id=bill_of_materials_yield_id
    )

    valid_fields = [
        "quantity_weight",
        "unit_weight",
        "quantity_volume",
        "unit_volume",
        "quantity_each",
        "note_each",
        "unit_each",
        "scale_multiple_smallest",
        "scale_multiple_largest",
    ]

    for field in valid_fields:
        if field in data:
            setattr(some_object, field, data[field])

    if data["quantity_weight"] and data["unit_weight"]:
        quantity_weight_standard, unit_weight_standard = standardize_units(
            quantity=data["quantity_weight"], unit=data["unit_weight"]
        )

    if data["quantity_volume"] and data["unit_volume"]:
        quantity_volume_standard, unit_volume_standard = standardize_units(
            quantity=data["quantity_volume"], unit=data["unit_volume"]
        )

    yields.full_clean()
    yields.save()

    return yields


def calculate_scaling_factor_for_bom(*, bom_line: models.BillOfMaterialsLine) -> D:

    # TODO: Standardize unit/quantity
    if bom_line.unit.unit_type == "WEIGHT":
        line_base_yield_quantity = (
            bom_line.item.get_bill_of_materials().yields.quantity_weight
        )
        line_base_yield_unit = bom_line.item.get_bill_of_materials().yields.unit_weight

    elif bom_line.unit.unit_type == "VOLUME":
        line_base_yield_quantity = (
            bom_line.item.get_bill_of_materials().yields.quantity_volume
        )
        line_base_yield_unit = bom_line.item.get_bill_of_materials().yields.unit_volume
    else:
        line_base_yield_quantity = (
            bom_line.item.get_bill_of_materials().yields.quantity_each
        )
        line_base_yield_unit = bom_line.item.get_bill_of_materials().yields.unit_each

    (standardized_base_quantity, standardized_base_unit) = standardize_units(
        quantity=line_base_yield_quantity, unit=line_base_yield_unit
    )
    (standardized_parent_quantity, standardized_parent_unit) = standardize_units(
        quantity=bom_line.quantity, unit=bom_line.unit
    )

    return standardized_parent_quantity / standardized_base_quantity


def scale_bom_lines_by_scaling_factor(lines, scaling_factor):
    """ Scale BOM line items by scaling factor """

    from apps.masterdata.services import adjust_units_for_display

    scaled_lines = []

    for line in lines:

        scaled_quantity = line.quantity * D(scaling_factor)

        display_quantity, display_unit = adjust_units_for_display(
            scaled_quantity, line.unit
        )

        scaled_lines.append(
            {
                "sequence": line.sequence,
                "quantity": display_quantity,
                "unit": display_unit,
                "item": line.item,
                "note": line.note,
            }
        )

    return scaled_lines


def scale_bom_resources_by_scaling_factor(resources, scaling_factor):
    """ Scale BOM resources required by scaling factor """

    scaled_resources = []

    for resource in resources:

        scaled_capacity_required = math.ceil(
            resource.capacity_required * D(scaling_factor)
        )

        scaled_resources.append(
            {
                "sequence": resource.sequence,
                "capacity_required": scaled_capacity_required,
                "resource": resource.resource,
                "note": resource.note,
            }
        )

    return scaled_resources


def scale_bom_yields_by_scaling_factor(yields, scaling_factor):
    """ Scale BOM yields by scaling factor """

    from apps.masterdata.services import adjust_units_for_display

    scaled_yield = {
        "quantity_weight": None,
        "unit_weight": None,
        "quantity_volume": None,
        "unit_volume": None,
        "quantity_each": None,
        "unit_each": None,
        "note_each": None,
    }

    if yields.quantity_weight and yields.unit_weight:
        scaled_quantity = yields.quantity_weight * D(scaling_factor)
        display_quantity, display_unit = adjust_units_for_display(
            scaled_quantity, yields.unit_weight
        )
        scaled_yield["quantity_weight"] = display_quantity
        scaled_yield["unit_weight"] = display_unit

    if yields.quantity_volume and yields.unit_volume:
        scaled_quantity = yields.quantity_volume * D(scaling_factor)
        display_quantity, display_unit = adjust_units_for_display(
            scaled_quantity, yields.unit_volume
        )
        scaled_yield["quantity_volume"] = display_quantity
        scaled_yield["unit_volume"] = display_unit

    if all(
        [
            yields.quantity_each,
            yields.unit_each,
            yields.note_each,
        ]
    ):
        scaled_quantity = yields.quantity_each * D(scaling_factor)
        scaled_yield["quantity_each"] = scaled_quantity
        scaled_yield["unit_each"] = yields.unit_each
        scaled_yield["note_each"] = yields.note_each

    return scaled_yield


def scale_bill_of_materials_by_multiple(
    bill_of_materials,
    scaling_factor,
):
    """ Accepts BOM and scaling factor, returns scaled materials and yields for displaying/printing/exporting scaled BOM """

    from apps.masterdata.services import (
        adjust_units_for_display,
        scale_bom_lines_by_scaling_factor,
        scale_bom_resources_by_scaling_factor,
        scale_bom_yields_by_scaling_factor,
    )

    """ use user-specified scaling_factor to adjust lines, resources & yields """

    scaled_lines = scale_bom_lines_by_scaling_factor(
        bill_of_materials.lines.all(), scaling_factor
    )
    scaled_resources = scale_bom_resources_by_scaling_factor(
        bill_of_materials.resource_requirements.all(), scaling_factor
    )
    scaled_yields = scale_bom_yields_by_scaling_factor(
        bill_of_materials.yields, scaling_factor
    )

    return scaled_lines, scaled_resources, scaled_yields


def scale_bill_of_materials_by_yield(
    bill_of_materials,
    desired_quantity,
    desired_unit,
    # yield_unit_type,
):
    """Accepts BOM and parameters of desired yield, calculates scaling factor from these, and
    returns scaled materials and yields for displaying/printing/exporting scaled BOM"""

    from apps.masterdata.services import (
        adjust_units_for_display,
        convert_units,
        scale_bom_lines_by_scaling_factor,
        scale_bom_resources_by_scaling_factor,
        scale_bom_yields_by_scaling_factor,
    )

    """ calculate scaling factor """
    yield_unit_type = desired_unit.unit_type
    yield_unit_type_property_quantity = f"quantity_{yield_unit_type.lower()}"
    batch_yield_quantity = getattr(
        bill_of_materials.yields, yield_unit_type_property_quantity
    )

    yield_unit_type_property_unit = f"unit_{yield_unit_type.lower()}"
    batch_yield_unit = getattr(bill_of_materials.yields, yield_unit_type_property_unit)

    if yield_unit_type == "EACH":
        scaling_factor = desired_quantity / batch_yield_quantity
    else:
        standardized_batch_yield_quantity = convert_units(
            D(batch_yield_quantity),
            batch_yield_unit.symbol,
            STANDARD_UNITS[yield_unit_type],
        )
        standardized_desired_yield_quantity = convert_units(
            D(desired_quantity),
            desired_unit.symbol,
            STANDARD_UNITS[yield_unit_type],
        )
        scaling_factor = (
            standardized_desired_yield_quantity / standardized_batch_yield_quantity
        )

    """ use calculated scaling_factor to adjust lines, resources & yields """

    scaled_lines = scale_bom_lines_by_scaling_factor(
        bill_of_materials.lines.all(), scaling_factor
    )
    scaled_resources = scale_bom_resources_by_scaling_factor(
        bill_of_materials.resource_requirements.all(), scaling_factor
    )
    scaled_yields = scale_bom_yields_by_scaling_factor(
        bill_of_materials.yields, scaling_factor
    )

    return scaled_lines, scaled_resources, scaled_yields, round(scaling_factor, 2)


def scale_bill_of_materials_by_limit(
    bill_of_materials,
    limiting_quantity,
    limiting_unit,
    limiting_line,
):
    """Accepts BOM and parameters of limiting line item, calculates scaling factor from these, and
    returns scaled materials and yields for displaying/printing/exporting scaled BOM"""

    from apps.masterdata.services import (
        adjust_units_for_display,
        convert_units,
        scale_bom_lines_by_scaling_factor,
        scale_bom_resources_by_scaling_factor,
        scale_bom_yields_by_scaling_factor,
    )

    """ calculate scaling factor """

    # limiting_line = bill_of_materials.lines.get(item_id=limiting_item_id)

    batch_line_quantity = limiting_line.quantity
    batch_line_unit = limiting_line.unit

    if batch_line_unit.unit_type == "EACH":
        scaling_factor = limiting_quantity / batch_line_quantity
    else:
        standardized_batch_line_quantity = convert_units(
            D(batch_line_quantity),
            batch_line_unit.symbol,
            STANDARD_UNITS[batch_line_unit.unit_type],
        )
        standardized_limiting_line_quantity = convert_units(
            D(limiting_quantity),
            limiting_unit.symbol,
            STANDARD_UNITS[limiting_unit.unit_type],
        )
        scaling_factor = (
            standardized_limiting_line_quantity / standardized_batch_line_quantity
        )

    """ use calculated scaling_factor to adjust lines, resource & yields """

    scaled_lines = scale_bom_lines_by_scaling_factor(
        bill_of_materials.lines.all(), scaling_factor
    )
    scaled_resources = scale_bom_resources_by_scaling_factor(
        bill_of_materials.resource_requirements.all(), scaling_factor
    )
    scaled_yields = scale_bom_yields_by_scaling_factor(
        bill_of_materials.yields, scaling_factor
    )

    return scaled_lines, scaled_resources, scaled_yields, round(scaling_factor, 2)


def store_scaling_variables_in_session(request):
    if request.POST.get("scale_type") == "YIELD":
        request.session["scale_type"] = "YIELD"
        yield_choice = request.POST.get("yield_choice")
        request.session["quantity"] = request.POST.get(
            f"yield_quantity_{yield_choice.lower()}"
        )
        request.session["unit"] = request.POST.get(f"yield_unit_{yield_choice.lower()}")

    if request.POST.get("scale_type") == "MULTIPLE":
        request.session["scale_type"] = "MULTIPLE"
        request.session["scaling_factor"] = request.POST.get("multiple_scaling_factor")

    if request.POST.get("scale_type") == "LIMIT":
        request.session["scale_type"] = "LIMIT"
        request.session["quantity"] = request.POST.get("limit_quantity")
        request.session["unit"] = request.POST.get("limit_unit")
        request.session["line"] = request.POST.get("limit_line")

    return


def generate_bom_tree(*, root_bom: models.BillOfMaterials) -> None:
    """New implementation using Treebeard to persist n-ary tree once built
    Tree will be able for lookup rather than being re-built for each production view
    Tree will be built upon BOM transition to approved"""

    PRODUCT_CHOICES = ["WIP", "FINISHED"]

    node_lookup = {}
    offsets = []
    queue = deque([])
    visited_nodes = []

    get = lambda node_id: models.BillOfMaterialsTree.objects.get(pk=node_id)

    root = models.BillOfMaterialsTree.add_root(
        name=root_bom.product.name,
        scaling_factor=1,
        bill_of_materials=root_bom,
        node_bom=root_bom,
    )
    node_lookup[root_bom.id] = root
    queue.extend(list(root_bom.lines.all()))

    counter = 0
    while len(queue) > 0:

        flag = 0  # checks whether all the child nodes have been visited.

        # if next element of queue is Material (leaf node), remove from queue:
        if queue[-1].item.category not in PRODUCT_CHOICES:
            queue.pop()
            continue
        # if next element of queue is has already been visited:
        elif queue[-1].id in visited_nodes:
            queue.pop()
            continue
        else:
            current = queue[-1]
            parent_node = node_lookup[current.bill_of_materials_id]

            scaling_factor = (
                calculate_scaling_factor_for_bom(bom_line=current)
                * parent_node.scaling_factor
            )

            # (standardized_quantity, standardized_unit) = standardize_units(
            #     quantity=current.quantity, unit=current.unit
            # )

            current_node = get(parent_node.id).add_child(
                name=current.item.get_bill_of_materials().product.name,
                scaling_factor=scaling_factor,
                leadtime=current.item.get_bill_of_materials().characteristics.leadtime,
                # quantity=standardized_quantity,
                quantity_standard=current.quantity_standard,
                # unit=standardized_unit,
                unit_standard=current.unit_standard,
                bill_of_materials=None,
                node_bom=current.item.get_bill_of_materials(),
            )

            node_lookup[current.item.get_bill_of_materials().id] = current_node

        # any unvisited child (left to right sequence), is pushed to queue and added to visited_nodes
        counter = 0
        for line in current.item.get_bill_of_materials().lines.all():
            if line.item.category in PRODUCT_CHOICES:
                flag = 1
                queue.append(line)
                visited_nodes.append(current.id)
                counter += 1
            else:
                visited_nodes.append(current.id)

        # if all child nodes (from left to right) of current (parent node) have been visited then remove parent node from queue
        if flag == 0:
            queue.pop()
        counter += 1

    return root


def generate_cost_sequence(bom_obj):
    # returns a ordered list of sub-costings to calculate

    product_choices = ["WIP", "FINISHED"]

    evaluation_queue = [bom_obj]  # evaluation_queue must be only BOMs
    return_data = []  # return_data will contain only BOMs
    return_errors = []

    while len(evaluation_queue):
        current = evaluation_queue.pop()
        if current.product.category in product_choices:
            return_data.insert(0, current)
            for line in current.lines.all():
                if line.item.category in product_choices:
                    evaluation_queue.append(line.item.bills_of_materials.latest())
            continue

        else:
            return_data.insert(0, current)
            for line in current.lines.all():
                if line.item.category in product_choices:
                    evaluation_queue.append(line.item.bills_of_materials.latest())
            continue
    return return_data


def calculate_cost_sequence(seq):
    product_choices = ["WIP", "FINISHED"]
    threshold_for_cost_drivers = D(0.20)

    cost_lookup = {}
    cost_table = []

    for idx, bom in enumerate(seq):
        total_cost = D(0)
        temp = {"item": bom.product.name}
        lines = []

        for line in bom.lines.all():

            # if mismatch of units
            # try item_conversion

            if line.item.category not in product_choices:
                # convert quantity to standard units
                if line.item.unit_type == "EACH":
                    standardized_quantity = line.quantity
                    unit_cost = line.item.costs.latest().unit_cost_each
                elif line.item.unit_type == "WEIGHT":
                    standardized_quantity = convert_units(
                        D(line.quantity),
                        line.unit.symbol,
                        STANDARD_UNITS[line.item.unit_type],
                    )
                    unit_cost = line.item.costs.latest().unit_cost_weight
                else:
                    standardized_quantity = convert_units(
                        D(line.quantity),
                        line.unit.symbol,
                        STANDARD_UNITS[line.item.unit_type],
                    )
                    unit_cost = line.item.costs.latest().unit_cost_volume

                # calculate extension
                extension = round(standardized_quantity * unit_cost, 3)
                total_cost += D(extension)

            else:
                # convert quantity to standard units
                # determine scaling factor = bom quantity/batch yield
                if line.item.unit_type == "EACH":
                    standardized_quantity = line.quantity
                    scaling_factor = (
                        line.quantity
                        / line.item.bills_of_materials.latest().yields.quantity_each
                    )
                elif line.item.unit_type == "WEIGHT":
                    standardized_quantity = convert_units(
                        D(line.quantity),
                        line.unit.symbol,
                        STANDARD_UNITS["WEIGHT"],
                    )
                    standardized_yield = convert_units(
                        D(line.item.bills_of_materials.latest().yields.quantity_weight),
                        line.item.bills_of_materials.latest().yields.unit_weight.symbol,
                        STANDARD_UNITS["WEIGHT"],
                    )
                    scaling_factor = standardized_quantity / standardized_yield
                else:
                    standardized_quantity = convert_units(
                        D(line.quantity),
                        line.unit.symbol,
                        STANDARD_UNITS["VOLUME"],
                    )
                    standardized_yield = convert_units(
                        D(line.item.bills_of_materials.latest().yields.quantity_volume),
                        line.item.bills_of_materials.latest().yields.unit_volume.symbol,
                        STANDARD_UNITS["VOLUME"],
                    )
                    scaling_factor = standardized_quantity / standardized_yield

                # TODO: Need to figure out unit cost here
                unit_cost = D(0)

                # extension is batch cost in lookup * scaling factor
                extension = round(
                    cost_lookup.get(line.item.bills_of_materials.latest().id, 0)
                    * scaling_factor,
                    3,
                )
                total_cost += D(extension)

            lines.append(
                {
                    "sequence": line.sequence,
                    "quantity": line.quantity,
                    "unit": line.unit.symbol,
                    "item": line.item.name,
                    "id": line.item.id,
                    "standardized_quantity": standardized_quantity,
                    "standardized_unit": STANDARD_UNITS[line.item.unit_type],
                    "unit_cost": unit_cost,
                    "extension": extension,
                    "cost_driver": False,
                }
            )

        for line in lines:
            if line["extension"] > total_cost * threshold_for_cost_drivers:
                line["cost_driver"] = True

        temp["lines"] = lines
        cost_table.append(temp)
        cost_lookup[bom.id] = total_cost

    return cost_lookup, cost_table


def calculate_tree_cost(*, sequence: list) -> Tuple[dict, list]:
    product_choices = ["WIP", "FINISHED"]
    threshold_for_cost_drivers = D(0.20)

    cost_lookup = {}
    cost_table = []

    # calculate cost
    for idx, node in enumerate(sequence):
        bom = node.node_bom
        total_cost = D(0)
        temp = {"item": bom.product.name}
        lines = []

        for line in bom.lines.all():

            # if mismatch of units
            # try item_conversion

            if line.item.category not in product_choices:
                # convert quantity to standard units
                if line.item.unit_type == "EACH":
                    standardized_quantity = line.quantity
                    unit_cost = line.item.costs.latest().unit_cost_each
                elif line.item.unit_type == "WEIGHT":
                    standardized_quantity = convert_units(
                        D(line.quantity),
                        line.unit.symbol,
                        STANDARD_UNITS[line.item.unit_type],
                    )
                    unit_cost = line.item.costs.latest().unit_cost_weight
                else:
                    standardized_quantity = convert_units(
                        D(line.quantity),
                        line.unit.symbol,
                        STANDARD_UNITS[line.item.unit_type],
                    )
                    unit_cost = line.item.costs.latest().unit_cost_volume

                # calculate extension
                extension = round(standardized_quantity * unit_cost, 3)
                total_cost += D(extension)

            else:
                # convert quantity to standard units
                # determine scaling factor = bom quantity/batch yield
                if line.item.unit_type == "EACH":
                    standardized_quantity = line.quantity
                    scaling_factor = (
                        line.quantity
                        / line.item.bills_of_materials.latest().yields.quantity_each
                    )
                elif line.item.unit_type == "WEIGHT":
                    standardized_quantity = convert_units(
                        D(line.quantity),
                        line.unit.symbol,
                        STANDARD_UNITS["WEIGHT"],
                    )
                    standardized_yield = convert_units(
                        D(line.item.bills_of_materials.latest().yields.quantity_weight),
                        line.item.bills_of_materials.latest().yields.unit_weight.symbol,
                        STANDARD_UNITS["WEIGHT"],
                    )
                    scaling_factor = standardized_quantity / standardized_yield
                else:
                    standardized_quantity = convert_units(
                        D(line.quantity),
                        line.unit.symbol,
                        STANDARD_UNITS["VOLUME"],
                    )
                    standardized_yield = convert_units(
                        D(line.item.bills_of_materials.latest().yields.quantity_volume),
                        line.item.bills_of_materials.latest().yields.unit_volume.symbol,
                        STANDARD_UNITS["VOLUME"],
                    )
                    scaling_factor = standardized_quantity / standardized_yield

                # TODO: Need to figure out unit cost here
                unit_cost = D(0)

                # extension is batch cost in lookup * scaling factor
                extension = round(
                    cost_lookup.get(line.item.bills_of_materials.latest().id, 0)
                    * scaling_factor,
                    3,
                )
                total_cost += D(extension)

            lines.append(
                {
                    "sequence": line.sequence,
                    "quantity": line.quantity,
                    "unit": line.unit.symbol,
                    "item": line.item.name,
                    "id": line.item.id,
                    "standardized_quantity": standardized_quantity,
                    "standardized_unit": STANDARD_UNITS[line.item.unit_type],
                    "unit_cost": unit_cost,
                    "extension": extension,
                    "cost_driver": False,
                }
            )

        for line in lines:
            if line["extension"] > total_cost * threshold_for_cost_drivers:
                line["cost_driver"] = True

        temp["lines"] = lines

        if idx == len(sequence) - 1:
            cost_table.append(temp)
        cost_lookup[bom.id] = total_cost

    return cost_lookup, json.dumps(cost_table, cls=DecimalEncoder)


""" 
WeasyPrint currently (v.52.5, as of 2021-5-4) has not yet implemented CSS Grid  
- so the html template used for rendering for OrderDetails for export has been rewritten using tables
- there may be some inconsistencies in rendering, have to keep checking
"""

REPORT_TEMPLATES = {
    "brief": "reports/bill_of_materials_pdf_brief.html",
    "standard": "reports/bill_of_materials_pdf_standard.html",
    "full": "reports/bill_of_materials_pdf_full.html",
}


def generate_pdf_for_bill_of_materials(
    *,
    base_url: str,
    bill_of_materials_id: str,
    report_type: str,
    view_type: str,
    organization: str,
    scale_type: str = None,
    scaling_factor: decimal.Decimal = None,
    quantity: decimal.Decimal = None,
    unit: models.UnitMeasurement = None,
    line: models.BillOfMaterialsLine = None,
) -> HttpResponse:

    # Retrieve/calculate context
    printed_as_of = f"Printed {timezone.now().astimezone(EASTERN_TZ).strftime('%A, %B %d, %Y at %I:%M %p')}"

    data = models.BillOfMaterials.objects.get(id=bill_of_materials_id)
    options = common_services.get_template_context_options(models.BillOfMaterials)

    if view_type == "scaled":
        # scale by multiple
        if scale_type == "MULTIPLE" and all([data, scaling_factor]):
            (
                scaled_lines,
                scaled_resources,
                scaled_yield,
            ) = scale_bill_of_materials_by_multiple(data, scaling_factor)
            scaling_description = f"Scaled {'up' if D(scaling_factor) > 1 else 'down'} to <strong>{scaling_factor}<em>x</em></strong> standard batch"

        # scale by yield
        if scale_type == "YIELD" and all([data, quantity, unit]):
            (
                scaled_lines,
                scaled_resources,
                scaled_yield,
                scaling_factor,
            ) = scale_bill_of_materials_by_yield(data, quantity, unit)
            scaling_description = f"Scaled {'up' if scaling_factor > 1 else 'down'} to produce <strong>{quantity} {unit.symbol} {data.yields.yield_noun or ''} {data.yields.note_each or '' if unit.unit_type == 'EACH' else ''}</strong> ({scaling_factor}<em>x</em> standard batch)"

        # scale by limiting line
        if scale_type == "LIMIT" and all([data, quantity, unit, line]):
            (
                scaled_lines,
                scaled_resources,
                scaled_yield,
                scaling_factor,
            ) = services.scale_bill_of_materials_by_limit(data, quantity, unit, line)
            scaling_description = f"Scaled {'up' if scaling_factor > 1 else 'down'} to use <strong>{round(quantity,0) if line.unit.symbol in ['g', 'ml'] else round(quantity, 3)} {unit.symbol} {line.item.name.title()}</strong> ({scaling_factor}<em>x</em> standard batch)"

        scaling_factor = str(scaling_factor)
        lines = scaled_lines
        resources = scaled_resources
        yields = scaled_yield

    else:
        scaling_factor = None
        scaling_description = None
        lines = data.lines.all()
        resources = data.resource_requirements.all()
        yields = data.yields

    header = f"{data.product.name} (v.{data.version} {data.get_state_display().upper()}{ scaling_factor if view_type == 'scaled' else ''}{'x scale' if view_type == 'scaled' else ''})"
    header = header.replace('"', '\\"')

    # Render PDF
    font_config = FontConfiguration()
    template = REPORT_TEMPLATES.get(report_type, None)

    html_string = render_to_string(
        template,
        {
            "data": data,
            "options": options,
            "view_type": view_type,
            "scaling_factor": scaling_factor,
            "scaling_description": scaling_description,
            "lines": lines,
            "resources": resources,
            "yields": yields,
            "organization": organization,
            "printed_as_of": printed_as_of,
            "header": header,
        },
    )

    html = HTML(string=html_string, base_url=base_url)
    result = html.write_pdf(
        font_config=font_config,
        presentational_hints=True,
    )

    # Generate HTTP response
    filename = f"{data.product.name}-v.{data.version}.pdf"

    response = HttpResponse(content_type="application/pdf;")
    response["Content-Transfer-Encoding"] = "binary"
    response["Content-Disposition"] = f'inline;filename="{filename}"'
    # response["Content-Disposition"] = f'attachment;filename="{filename}"'

    with NamedTemporaryFile(delete=True) as output:
        output.write(result)
        output.flush()
        output = open(output.name, "rb")
        response.write(output.read())

    return response


# –––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––
# MATERIAL
# –––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––


# def material_get_assigned_cost(*, material: models.Material) -> str:
#     if self.costs.exists():
#         if self.unit_type == "WEIGHT":
#             return f"{self.costs.latest().unit_cost_weight}"
#         elif self.unit_type == "VOLUME":
#             return f"{self.costs.latest().unit_cost_volume}"
#         elif self.unit_type == "EACH":
#             return f"{self.costs.latest().unit_cost_each}"
#         else:
#             return f"{D(0.00)}"
#     else:
#         return "No costs entered yet"


def import_xlsx(
    *,
    model: Model,
    file: bin,
) -> dict:
    wb = load_workbook(filename=file.file)
    ws = wb.active

    # parse xlsx
    row_count = ws.max_row
    column_count = ws.max_column

    header = [cell.value for cell in ws[1]]

    objects = []
    for row in list(ws.rows)[1:]:
        temp = {}
        for key, cell in zip(header, row):
            temp[key] = cell.value
        objects.append(temp)

    # compare with database
    updates = []
    creates = []
    no_changes = []

    for item in objects:
        if item["id"] is None:
            creates.append(item)
        else:
            qs = model.objects.all()
            if qs.filter(**item).exists():
                no_changes.append(item)
            else:
                updates.append(item)

    data = {"updates": updates, "creates": creates, "no_changes": no_changes}

    return data


def export_as_xlsx_with_choices_validation(
    *,
    model: Model,
    queryset: QuerySet,
    fields_to_export: list,
    fields_to_validate: list,
    filename: str = "default",
) -> HttpResponse:
    DATA_ROW_RANGE = (2, 1048576)  # upper bound of rows in XLSX document

    # check that fields_to_validate is subset of fields_to_export
    if not set(fields_to_validate).issubset(set(fields_to_export)):
        return

    # define workbook
    wb = Workbook()

    # define data sheet
    data_sheet = wb.active
    data_sheet.title = model._meta.verbose_name_plural.title()

    # define columns and write header row
    columns = fields_to_export
    # columns = [
    #     model._meta.get_field(field_name).verbose_name
    #     for field_name in fields_to_export
    # ]
    row_number = 1
    for column_number, column_title in enumerate(columns, 1):
        cell = data_sheet.cell(row=row_number, column=column_number)
        cell.value = column_title

    # for each item in queryset, write row
    for obj in queryset:
        row_number += 1
        row = [getattr(obj, field_name) for field_name in fields_to_export]

        for column_number, cell_value in enumerate(row, 1):
            cell = data_sheet.cell(row=row_number, column=column_number)
            cell.value = cell_value

    # as user will be editing/adding to this sheet and uploading later,
    # apply validation to fields with choices defined

    # create sheet to hold values to validate against
    validation_sheet = wb.create_sheet(title=f"Validation - DO NOT MODIFY")
    max_row = 1

    cell = validation_sheet.cell(row=max_row, column=1)
    cell.value = "DO NOT MODIFY THIS SHEET"
    max_row += 1
    cell = validation_sheet.cell(row=max_row, column=1)
    cell.value = " "
    max_row += 1

    for field in fields_to_validate:

        # define columns in data sheet to apply validation to for this field
        field_column = fields_to_export.index(field) + 1
        cells_to_validate_range = CellRange(
            min_col=field_column,
            min_row=DATA_ROW_RANGE[0],
            max_col=field_column,
            max_row=DATA_ROW_RANGE[1],
        )

        # choices = [choice[0] for choice in model._meta.get_field(field).choices]
        # for each item in choices, write row on validation_sheet
        field_value_column_number = 1
        display_value_column_number = 2

        cell = validation_sheet.cell(row=max_row, column=field_value_column_number)
        cell.value = f"{model._meta.get_field(field).verbose_name}"
        max_row += 1

        initial_row_for_current_field = max_row
        max_row_for_current_field = max_row

        for choice in model._meta.get_field(field).choices:
            # choice field value
            cell = validation_sheet.cell(
                row=max_row_for_current_field, column=field_value_column_number
            )
            cell.value = choice[0]
            # choice display value
            cell = validation_sheet.cell(
                row=max_row_for_current_field, column=display_value_column_number
            )
            cell.value = choice[1]
            max_row_for_current_field += 1
            max_row += 1

        cell = validation_sheet.cell(row=max_row + 1, column=field_value_column_number)
        cell.value = f" "
        max_row += 1

        valid_values_range = CellRange(
            min_col=field_value_column_number,
            min_row=initial_row_for_current_field,
            max_col=field_value_column_number,
            max_row=max_row_for_current_field - 1,
        )
        valid_values_range_absolute = absolute_coordinate(valid_values_range.coord)

        dv = DataValidation(
            type="list",
            formula1=f"{quote_sheetname(validation_sheet.title)}!{valid_values_range_absolute}",
            allow_blank=True,
            error="Your entry is not an permitted choice",
            errorTitle="Invalid Entry",
        )
        dv.add(cells_to_validate_range)
        data_sheet.add_data_validation(dv)

    wb.active = data_sheet
    validation_sheet.protection.sheet = True
    validation_sheet.protection.enable()

    # wb.save(filename)

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()

    response = HttpResponse(
        content=stream,
        content_type="application/ms-excel",
    )
    response[
        "Content-Disposition"
    ] = f'attachment; filename={filename}-{dt.datetime.now().strftime("%Y%m%d%H%M")}.xlsx'

    return response


# def export_as_xlsx_with_choices_validation(
#     *,
#     model: Model,
#     queryset: QuerySet,
#     fields_to_export: list,
#     fields_to_validate: list,
#     filename: str = "default",
# ) -> HttpResponse:
#     DATA_ROW_RANGE = (2, 1048576)  # upper bound of rows in XLSX document

#     # check that fields_to_validate is subset of fields_to_export
#     if set(fields_to_validate).issubset(set(fields_to_export)):
#         pass
#     else:
#         return

#     # define workbook
#     wb = Workbook()

#     # define data sheet
#     data_sheet = wb.active
#     data_sheet.title = model._meta.verbose_name_plural.title()

#     # define columns and write header row
#     columns = [
#         model._meta.get_field(field_name).verbose_name
#         for field_name in fields_to_export
#     ]
#     row_number = 1
#     for column_number, column_title in enumerate(columns, 1):
#         cell = data_sheet.cell(row=row_number, column=column_number)
#         cell.value = column_title

#     # for each item in queryset, write row
#     for obj in queryset:
#         row_number += 1
#         row = [getattr(obj, field_name) for field_name in fields_to_export]

#         for column_number, cell_value in enumerate(row, 1):
#             cell = data_sheet.cell(row=row_number, column=column_number)
#             cell.value = cell_value

#     # as user will be editing/adding to this sheet and uploading later,
#     # apply validation to fields with choices defined
#     for field in fields_to_validate:

#         # define columns in data sheet to apply validation to for this field
#         field_column = fields_to_export.index(field) + 1

#         cells_to_validate_range = CellRange(
#             min_col=field_column,
#             min_row=DATA_ROW_RANGE[0],
#             max_col=field_column,
#             max_row=DATA_ROW_RANGE[1],
#         )

#         # create sheet to hold values to validate against
#         validation_sheet = wb.create_sheet(title=f"Validation {field}")
#         validation_sheet_name = f"Validation {field}"

#         choices = [choice[0] for choice in model._meta.get_field(field).choices]

#         # for each item in choices, write row on validation_sheet
#         cell = validation_sheet.cell(row=1, column=1)
#         cell.value = "DO NOT MODIFY THIS SHEET"

#         validation_row_number = 2
#         column_number = 1
#         for choice in choices:
#             cell = validation_sheet.cell(
#                 row=validation_row_number, column=column_number
#             )
#             cell.value = choice
#             validation_row_number += 1

#         valid_values_range = CellRange(
#             min_col=column_number,
#             min_row=2,
#             max_col=column_number,
#             max_row=validation_row_number,
#         )
#         valid_values_range_absolute = absolute_coordinate(valid_values_range.coord)

#         dv = DataValidation(
#             type="list",
#             formula1=f"{quote_sheetname(validation_sheet_name)}!{valid_values_range_absolute}",
#             allow_blank=True,
#             error="Your entry is not an permitted choice",
#             errorTitle="Invalid Entry",
#         )
#         dv.add(cells_to_validate_range)
#         data_sheet.add_data_validation(dv)

#     wb.save(filename)
#     return


def export_as_csv(
    *, queryset: QuerySet, fields: list, filename: str = "default"
) -> HttpResponse:
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename={}.csv".format(filename)
    writer = csv.writer(response)

    writer.writerow(fields)
    for obj in queryset:
        row = writer.writerow([getattr(obj, field) for field in fields])

    return response


def export_as_xlsx(
    *, queryset: QuerySet[models.Material], fields_to_export: list, filename: str
) -> HttpResponse:
    pass


def export_as_pdf(
    *, queryset: QuerySet[models.Material], fields_to_export: list, filename: str
) -> HttpResponse:
    pass


def import_as_csv(
    *, queryset: QuerySet[models.Material], fields_to_export: list, filename: str
) -> HttpResponse:
    pass


def import_as_xlsx(
    *, queryset: QuerySet[models.Material], fields_to_export: list, filename: str
) -> HttpResponse:
    pass


# –––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––
# PRODUCT
# –––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––


def product_flatten_BOM_version_history(*, product: models.Product) -> None:
    """
    - When duplicating products, new product should retain only latest APPROVED BOM (or DRAFT BOM if only available)
    - Bill of materials and its child objects should be set to version 1
    - Since will want to make edits immediately to duplicate, duplicate BOM will begin in DRAFT state to permit editing
    """

    # if no BOMs, return
    if product.bills_of_materials.count() == 0:
        return

    # if single BOM, set as DRAFT and retain
    if product.bills_of_materials.count() == 1:
        bom = (
            product.bills_of_materials.get()
        )  # retrieves regardless of state, unlike get_bill_of_materials()
        bom.state = "DRAFT"
        bom.save()
        return

    # if multiple BOMs, keep latest approved, set as DRAFT and discard others
    if product.bills_of_materials.count() > 1:
        bom = (
            models.BillOfMaterials.objects.filter(product=product, state="APPROVED")
            .order_by("created_at")
            .first()
        )
        bom.version = 1
        bom.state = "DRAFT"
        bom.save()

        discards = product.bills_of_materials.exclude(id=bom.id)
        discards.delete()
        return


def product_duplicate(*, product: models.Product) -> models.Product:
    """
    Duplicate product and related reference data
    DO NOT duplicate transactional data (Sales, Production, Compliance)
    Append 'DUPLICATE' to name
    Reset version to 1 and status to DRAFT
    """

    from apps.masterdata.services import product_flatten_BOM_version_history

    product_duplicate = common_services.clone_object_instance(
        obj=product,
        attrs_to_apply={"version": 1},
        exclude_app_labels=[
            "compliance",
            "deliveries",
            "production",
            "purchasing",
            "sales",
        ],
        exclude_child_models=[
            "menu",
            "orderline",
            "request",
            "requestline",
            "shiftline",
            "standingorderline",
        ],
    )

    product_duplicate.name = f"{product.name} DUPLICATE"
    product_duplicate.save()

    product_flatten_BOM_version_history(product=product_duplicate)

    return product_duplicate


def product_generate_version(*, product: models.Product) -> models.Product:
    """
    Duplicate product and related reference data
    DO NOT duplicate transactional data (Sales, Production, Compliance)
    Increment version
    Set status to DRAFT
    """

    product_new_version = common_services.clone_object_instance(
        obj=product,
        attrs_to_apply={"version": product.version + 1},
        exclude_app_labels=[
            "compliance",
            "deliveries",
            "production",
            "purchasing",
            "sales",
        ],
        exclude_child_models=[
            "menu",
            "orderline",
            "request",
            "requestline",
            "shiftline",
            "standingorderline",
        ],
    )

    return product_new_version


# –––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––
# RESOURCE
# –––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––


# –––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––
# TEAM
# –––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––––
